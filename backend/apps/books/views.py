from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Count, Q, Avg
from django.core.paginator import Paginator
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils.crypto import get_random_string
from .models import Book, LoanRecord, Announcement, Category, SiteConfig, Reservation, Review, ReviewReply, BookList, BookListEntry, BorrowRule
from apps.users.models import User, CreditLog
from datetime import date, timedelta
from django.utils import timezone
import uuid
import json

# ... (Previous simple views: home, admin_dashboard)

@login_required
def admin_dashboard(request):
    if request.user.role != 'admin':
        return redirect('home')
    
    from django.db.models import Sum
    from datetime import datetime
    
    today = date.today()
    first_day_of_month = today.replace(day=1)
    
    pending_payments = LoanRecord.objects.filter(status='pending_payment')
    total_pending_fine = sum(loan.calculate_fine() for loan in pending_payments)
    
    paid_this_month = LoanRecord.objects.filter(
        payment_date__gte=first_day_of_month,
        fine_paid=True
    ).aggregate(total=Sum('fine_amount'))['total'] or 0
    
    stats = {
        'total_books': Book.objects.count(),
        'total_users': User.objects.count(),
        'active_loans': LoanRecord.objects.filter(status='borrowed').count(),
        'pending_requests': LoanRecord.objects.filter(status='pending').count(),
        'pending_payments': pending_payments.count(),
        'monthly_fine_total': float(paid_this_month) + total_pending_fine,
        'monthly_fine_paid': float(paid_this_month),
        'monthly_fine_unpaid': total_pending_fine,
    }
    
    recent_loans = LoanRecord.objects.all().order_by('-borrow_date')[:5]
    return render(request, 'admin/dashboard.html', {'stats': stats, 'recent_loans': recent_loans})

@login_required
def dashboard_chart_data(request):
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
        
    today = timezone.now().date()
    
    # Weekly Data: Last 7 days
    weekly_labels = []
    weekly_data = []
    for i in range(6, -1, -1):
        target_date = today - timedelta(days=i)
        weekly_labels.append(target_date.strftime('%m-%d'))
        count = LoanRecord.objects.filter(borrow_date=target_date).count()
        weekly_data.append(count)
        
    # Monthly Data: Last 4 weeks (simplified approach: 4 segments of 7 days)
    monthly_labels = ['前四周', '前三周', '前两周', '本周']
    monthly_data = []
    for i in range(3, -1, -1):
        start_date = today - timedelta(days=(i * 7) + 6)
        end_date = today - timedelta(days=i * 7)
        count = LoanRecord.objects.filter(borrow_date__range=[start_date, end_date]).count()
        monthly_data.append(count)
        
    return JsonResponse({
        'weekly': {
            'labels': weekly_labels,
            'data': weekly_data
        },
        'monthly': {
            'labels': monthly_labels,
            'data': monthly_data
        }
    })

@login_required
def book_manage(request):
    if request.user.role != 'admin':
        return redirect('home')
    
    check_expired_reservations()
    
    query = request.GET.get('q', '')
    books_list = Book.objects.all().order_by('-created_at')
    
    if query:
        books_list = books_list.filter(Q(title__icontains=query) | Q(author__icontains=query) | Q(isbn__icontains=query))
    
    paginator = Paginator(books_list, 10)
    page_number = request.GET.get('page')
    books = paginator.get_page(page_number)
    
    reservation_counts = {}
    for book in books:
        count = Reservation.objects.filter(book=book, status__in=['waiting', 'notified']).count()
        reservation_counts[book.id] = count
    
    top_categories = Category.objects.filter(parent=None).prefetch_related('children')
    return render(request, 'admin/book_list.html', {
        'books': books,
        'top_categories': top_categories,
        'reservation_counts': reservation_counts
    })

@login_required
def book_create(request):
    if request.user.role != 'admin':
        return redirect('home')
        
    if request.method == 'POST':
        title = request.POST.get('title')
        author = request.POST.get('author')
        isbn = request.POST.get('isbn')
        category_id = request.POST.get('category')
        description = request.POST.get('description')
        total_stock = int(request.POST.get('total_stock', 0))
        cover = request.FILES.get('cover')
        
        if Book.objects.filter(isbn=isbn).exists():
            messages.error(request, "ISBN 已存在，请检查输入。")
        else:
            category = Category.objects.get(pk=category_id) if category_id else None
            book = Book.objects.create(
                title=title,
                author=author,
                isbn=isbn,
                category=category,
                description=description,
                stock=total_stock, # Initial stock equals total stock
                total_stock=total_stock,
                cover=cover
            )
            from .utils import ensure_barcode_exists
            ensure_barcode_exists(book)
            messages.success(request, f"图书《{title}》已成功上架。")
    
    return redirect('book_manage')

@login_required
def book_edit(request, pk):
    if request.user.role != 'admin':
        return redirect('home')
        
    book = get_object_or_404(Book, pk=pk)
    old_stock = book.stock
    old_isbn = book.isbn
    
    if request.method == 'POST':
        book.title = request.POST.get('title')
        book.author = request.POST.get('author')
        # ISBN typically shouldn't be changed easily or needs validation, but allowing for correction
        new_isbn = request.POST.get('isbn')
        if new_isbn != book.isbn and Book.objects.filter(isbn=new_isbn).exists():
             messages.error(request, "新的 ISBN 已存在。")
             return redirect('book_manage')
             
        book.isbn = new_isbn
        category_id = request.POST.get('category')
        book.category = Category.objects.get(pk=category_id) if category_id else None
        book.description = request.POST.get('description')
        
        # Stock logic: Update total. If total increases, increase current stock.
        new_total = int(request.POST.get('total_stock', 0))
        diff = new_total - book.total_stock
        book.total_stock = new_total
        book.stock += diff
        
        if request.FILES.get('cover'):
            book.cover = request.FILES.get('cover')
            
        book.save()
        
        from .utils import ensure_barcode_exists, get_barcode_dir, normalize_isbn
        if new_isbn != old_isbn:
            old_normalized = normalize_isbn(old_isbn)
            if old_normalized:
                old_file = get_barcode_dir() / f"{old_normalized}.png"
                if old_file.exists():
                    old_file.unlink()
        ensure_barcode_exists(book)
        
        # If stock was 0 and now > 0, notify waiting readers
        if old_stock == 0 and book.stock > 0:
            notified_count = 0
            while book.stock > notified_count:
                notified = Reservation.notify_next_reader(book)
                if not notified:
                    break
                notified_count += 1
            if notified_count > 0:
                messages.info(request, f"已通知 {notified_count} 位预约读者图书到货。")
        
        messages.success(request, f"图书《{book.title}》信息已更新。")
        
    return redirect('book_manage')

@login_required
def book_delete(request, pk):
    if request.user.role != 'admin':
        return redirect('home')
    book = get_object_or_404(Book, pk=pk)
    book.delete()
    messages.success(request, f"图书《{book.title}》已成功删除。")
    return redirect('book_manage')

# ... (Keep existing loan_manage, book_browse, book_detail, borrow_request, my_loans, user_manage, audit_loan, system_settings, etc.)
@login_required
def loan_manage(request):
    if request.user.role != 'admin':
        return redirect('home')
    
    check_expired_reservations()
    
    loans_list = LoanRecord.objects.all().order_by('-borrow_date')
    
    # Filtering
    status = request.GET.get('status')
    user_query = request.GET.get('user')
    
    if status:
        loans_list = loans_list.filter(status=status)
    if user_query:
        loans_list = loans_list.filter(user__username__icontains=user_query)
        
    paginator = Paginator(loans_list, 10)
    page_number = request.GET.get('page')
    loans = paginator.get_page(page_number)
    
    reservation_counts = {}
    for loan in loans:
        count = Reservation.objects.filter(book=loan.book, status='waiting').count()
        reservation_counts[loan.book.id] = count
    
    return render(request, 'admin/loan_list.html', {
        'loans': loans,
        'reservation_counts': reservation_counts
    })

@login_required
def book_browse(request):
    query = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    parent_category_id = request.GET.get('parent_category', '')
    
    books = Book.objects.all().annotate(
        avg_rating=Avg('reviews__rating'),
        review_count=Count('reviews')
    )
    if query:
        books = books.filter(Q(title__icontains=query) | Q(author__icontains=query) | Q(isbn__icontains=query))
    if category_id:
        books = books.filter(category_id=category_id)
    elif parent_category_id:
        books = books.filter(category__parent_id=parent_category_id)
    
    top_categories = Category.objects.filter(parent=None).prefetch_related('children')
    return render(request, 'books/browse.html', {
        'books': books, 
        'top_categories': top_categories, 
        'query': query,
        'selected_category': category_id,
        'selected_parent': parent_category_id
    })

def get_book_rating_info(book):
    reviews = book.reviews.all()
    review_count = reviews.count()
    avg_rating = reviews.aggregate(Avg('rating'))['rating__avg']
    if avg_rating:
        avg_rating = round(avg_rating, 1)
    return {
        'review_count': review_count,
        'avg_rating': avg_rating,
        'full_stars': int(avg_rating) if avg_rating else 0,
        'has_half_star': avg_rating and (avg_rating - int(avg_rating)) >= 0.5
    }

def can_review_book(user, book):
    if not user.is_authenticated:
        return False
    return LoanRecord.objects.filter(
        user=user,
        book=book,
        status='returned'
    ).exists()

@login_required
def book_detail(request, pk):
    book = get_object_or_404(Book, pk=pk)
    user_reservation = None
    queue_count = 0
    
    rating_info = get_book_rating_info(book)
    reviews = book.reviews.all().select_related('user').prefetch_related('replies__user')
    can_review = can_review_book(request.user, book)
    
    has_reviewed = book.reviews.filter(user=request.user).exists() if request.user.is_authenticated else False
    
    active_rule = BorrowRule.get_active_rule()
    borrow_rule_summary = active_rule.get_rule_summary() if active_rule else "借期30天，可续借1次"
    
    from .utils import ensure_barcode_exists
    barcode_url = ensure_barcode_exists(book)
    
    if request.user.is_authenticated:
        user_reservation = Reservation.objects.filter(
            user=request.user,
            book=book,
            status__in=['waiting', 'notified']
        ).first()
        queue_count = Reservation.objects.filter(book=book, status='waiting').count()
    return render(request, 'books/detail.html', {
        'book': book,
        'user_reservation': user_reservation,
        'queue_count': queue_count,
        'rating_info': rating_info,
        'reviews': reviews,
        'can_review': can_review,
        'has_reviewed': has_reviewed,
        'borrow_rule_summary': borrow_rule_summary,
        'barcode_url': barcode_url
    })

@login_required
def review_create(request, pk):
    book = get_object_or_404(Book, pk=pk)
    
    if not can_review_book(request.user, book):
        messages.error(request, "只有借阅并归还过该书的读者才能发表评论。")
        return redirect('book_detail', pk=pk)
    
    if book.reviews.filter(user=request.user).exists():
        messages.error(request, "您已经对该书发表过评论了。")
        return redirect('book_detail', pk=pk)
    
    if request.method == 'POST':
        rating = int(request.POST.get('rating', 0))
        content = request.POST.get('content', '').strip()
        
        if rating < 1 or rating > 5:
            messages.error(request, "请选择有效的评分（1-5星）。")
        elif not content:
            messages.error(request, "评论内容不能为空。")
        elif len(content) > 500:
            messages.error(request, "评论内容不能超过500字。")
        else:
            Review.objects.create(
                book=book,
                user=request.user,
                rating=rating,
                content=content
            )
            messages.success(request, "评论发表成功！")
            return redirect('book_detail', pk=pk)
    
    return redirect('book_detail', pk=pk)

@login_required
def review_reply_create(request, review_pk):
    review = get_object_or_404(Review, pk=review_pk)
    
    if request.method == 'POST':
        content = request.POST.get('content', '').strip()
        
        if not content:
            messages.error(request, "回复内容不能为空。")
        elif len(content) > 500:
            messages.error(request, "回复内容不能超过500字。")
        else:
            ReviewReply.objects.create(
                review=review,
                user=request.user,
                content=content
            )
            messages.success(request, "回复成功！")
    
    return redirect('book_detail', pk=review.book.pk)

@login_required
def review_delete(request, pk):
    review = get_object_or_404(Review, pk=pk)
    book_pk = review.book.pk
    
    if request.user.role != 'admin' and review.user != request.user:
        messages.error(request, "您无权删除这条评论。")
        return redirect('book_detail', pk=book_pk)
    
    review.delete()
    messages.success(request, "评论及相关回复已删除。")
    
    from_param = request.GET.get('from')
    if from_param == 'manage':
        return redirect('review_manage')
    elif from_param == 'my_reviews':
        return redirect('my_reviews')
    return redirect('book_detail', pk=book_pk)

@login_required
def my_reviews(request):
    reviews = Review.objects.filter(user=request.user).select_related('book').order_by('-created_at')
    return render(request, 'user/my_reviews.html', {'reviews': reviews})

@login_required
def review_manage(request):
    if request.user.role != 'admin':
        return redirect('home')
    
    query = request.GET.get('q', '')
    reviews_list = Review.objects.all().select_related('user', 'book').prefetch_related('replies').order_by('-created_at')
    
    if query:
        reviews_list = reviews_list.filter(
            Q(content__icontains=query) | 
            Q(user__username__icontains=query) | 
            Q(book__title__icontains=query)
        )
    
    paginator = Paginator(reviews_list, 10)
    page_number = request.GET.get('page')
    reviews = paginator.get_page(page_number)
    
    return render(request, 'admin/review_manage.html', {
        'reviews': reviews,
        'query': query
    })

@login_required
def join_reservation(request, pk):
    book = get_object_or_404(Book, pk=pk)
    
    if book.stock > 0:
        messages.warning(request, "该书仍有库存，可直接借阅。")
        return redirect('book_detail', pk=pk)
    
    existing = Reservation.objects.filter(
        user=request.user,
        book=book,
        status__in=['waiting', 'notified']
    ).first()
    
    if existing:
        if existing.status == 'notified':
            messages.warning(request, "您已收到预约到货通知，请在48小时内发起借阅。")
        else:
            messages.warning(request, "您已在排队中。")
        return redirect('book_detail', pk=pk)
    
    Reservation.objects.create(
        user=request.user,
        book=book
    )
    messages.success(request, "成功加入预约队列！图书到货后将按顺序通知您。")
    return redirect('book_detail', pk=pk)

@login_required
def cancel_reservation(request, pk):
    reservation = get_object_or_404(Reservation, pk=pk, user=request.user)
    book = reservation.book
    
    if reservation.status in ['waiting', 'notified']:
        reservation.cancel()
        messages.success(request, "预约已取消。")
    
    return redirect('book_detail', pk=book.pk)

@login_required
def reservation_queue(request, pk):
    if request.user.role != 'admin':
        return redirect('home')
    book = get_object_or_404(Book, pk=pk)
    reservations = Reservation.objects.filter(book=book, status__in=['waiting', 'notified']).order_by('created_at')
    return render(request, 'admin/reservation_queue.html', {'book': book, 'reservations': reservations})

@login_required
def remove_reservation(request, pk):
    if request.user.role != 'admin':
        return redirect('home')
    reservation = get_object_or_404(Reservation, pk=pk)
    book_pk = reservation.book.pk
    reservation.cancel()
    messages.success(request, f"已移除 {reservation.user.username} 的预约。")
    return redirect('reservation_queue', pk=book_pk)

def check_expired_reservations():
    now = timezone.now()
    expired = Reservation.objects.filter(status='notified', expire_at__lte=now)
    for res in expired:
        res.status = 'expired'
        res.save()
        Reservation.update_queue_positions(res.book)
        Reservation.notify_next_reader(res.book)

@login_required
def borrow_request(request, pk):
    book = get_object_or_404(Book, pk=pk)
    
    check_expired_reservations()
    
    rule = BorrowRule.get_active_rule()
    
    if LoanRecord.objects.filter(user=request.user, status='pending_payment').exists():
        messages.error(request, "您存在未缴纳的逾期罚款，请先缴费后再申请借阅。")
        return redirect('my_loans')
    
    if book.stock <= 0:
        messages.error(request, "该图书目前无库存，无法借阅。")
        return redirect('book_detail', pk=pk)
        
    if not request.user.can_borrow():
        messages.error(request, f"信用不足，暂停借阅。当前信用分：{request.user.credit_score}，需≥60分方可借阅。")
        return redirect('book_detail', pk=pk)
        
    if LoanRecord.objects.filter(user=request.user, book=book, status__in=['pending', 'borrowed']).exists():
        messages.warning(request, "您已申请或正在借阅此书，请勿重复操作。")
        return redirect('book_detail', pk=pk)
    
    if rule and rule.max_borrow_quantity > 0:
        current_borrowed = LoanRecord.objects.filter(
            user=request.user, 
            status__in=['pending', 'borrowed']
        ).count()
        if current_borrowed >= rule.max_borrow_quantity:
            messages.error(request, f"借阅数量已达上限。当前规则：最多同时借阅{rule.max_borrow_quantity}本。")
            return redirect('book_detail', pk=pk)
    
    if rule and rule.max_daily_requests > 0:
        today_requests = LoanRecord.objects.filter(
            user=request.user, 
            borrow_date=date.today()
        ).count()
        if today_requests >= rule.max_daily_requests:
            messages.error(request, f"今日申请次数已达上限。当前规则：每日最多申请{rule.max_daily_requests}次。")
            return redirect('book_detail', pk=pk)
    
    reservation = Reservation.objects.filter(
        user=request.user,
        book=book,
        status='notified'
    ).first()
    
    config = SiteConfig.get_solo()
    
    borrow_days = rule.max_borrow_days if rule else 30
    rule_snapshot = rule.to_dict() if rule else {
        'name': '默认规则',
        'max_borrow_days': 30,
        'max_borrow_quantity': 0,
        'max_daily_requests': 0,
        'allow_renew': True,
        'max_renew_count': 1,
        'renew_days': 15,
    }
    
    import json
    LoanRecord.objects.create(
        user=request.user,
        book=book,
        due_date=date.today() + timedelta(days=borrow_days),
        status='pending',
        fine_daily_rate=config.daily_fine_rate,
        borrow_rule_snapshot=json.dumps(rule_snapshot)
    )
    
    if reservation:
        reservation.status = 'completed'
        reservation.save()
    
    messages.success(request, "借阅申请已提交，请等待管理员审核。")
    return redirect('my_loans')

@login_required
def my_loans(request):
    check_expired_reservations()
    loans = LoanRecord.objects.filter(user=request.user).order_by('-borrow_date')
    
    loans_with_fine = []
    for loan in loans:
        loan.current_fine = loan.calculate_fine()
        loan.overdue_days = loan.get_overdue_days()
        loans_with_fine.append(loan)
    
    reservations = Reservation.objects.filter(
        user=request.user,
        status__in=['waiting', 'notified']
    ).order_by('-created_at')
    return render(request, 'user/my_loans.html', {
        'loans': loans_with_fine,
        'reservations': reservations
    })

@login_required
@require_POST
def renew_loan(request, pk):
    loan = get_object_or_404(LoanRecord, pk=pk, user=request.user)
    
    if loan.status != 'borrowed':
        messages.error(request, "只有借阅中的记录才能续借。")
        return redirect('my_loans')
    
    if loan.is_overdue():
        messages.error(request, "该借阅已逾期，无法续借。请尽快归还图书。")
        return redirect('my_loans')
    
    if not loan.can_renew():
        rule = loan.get_borrow_rule()
        if not rule or not rule.get('allow_renew', False):
            messages.error(request, "该借阅规则不允许续借。")
        else:
            messages.error(request, "已达到最大续借次数。")
        return redirect('my_loans')
    
    rule = loan.get_borrow_rule()
    renew_days = rule.get('renew_days', 15)
    
    loan.due_date = loan.due_date + timedelta(days=renew_days)
    loan.renew_count += 1
    loan.save()
    
    messages.success(request, f"续借成功！应还日期已延长至 {loan.due_date.strftime('%Y-%m-%d')}。")
    return redirect('my_loans')
    
@login_required
def user_manage(request):
    if request.user.role != 'admin':
        return redirect('home')
    users = User.objects.all().exclude(pk=request.user.pk)
    return render(request, 'admin/user_list.html', {'users': users})

@login_required
def audit_loan(request, pk, action):
    if request.user.role != 'admin':
        return redirect('home')
    loan = get_object_or_404(LoanRecord, pk=pk)
    
    if action == 'approve':
        if loan.book.stock > 0:
            loan.status = 'borrowed'
            loan.book.stock -= 1
            loan.book.save()
            loan.save()
            
            reservation = Reservation.objects.filter(
                user=loan.user,
                book=loan.book,
                status='notified'
            ).first()
            if reservation:
                reservation.status = 'completed'
                reservation.save()
            
            if loan.book.stock > 0:
                Reservation.notify_next_reader(loan.book)
            
            messages.success(request, "借阅申请已批准。")
        else:
            messages.error(request, "库存不足，无法批准。")
    elif action == 'reject':
        loan.status = 'rejected'
        loan.save()
        messages.success(request, "借阅申请已拒绝。")
    elif action == 'return':
        loan.return_date = date.today()
        loan.book.stock += 1
        loan.book.save()
        
        days_diff = (loan.due_date - loan.return_date).days
        if days_diff >= 7:
            points = 3
            log_type = 'return_early'
            reason = f'提前归还《{loan.book.title}》，提前{days_diff}天'
            loan.status = 'returned'
            loan.fine_paid = True
        elif days_diff >= 0:
            points = 1
            log_type = 'return_on_time'
            reason = f'按时归还《{loan.book.title}》'
            loan.status = 'returned'
            loan.fine_paid = True
        else:
            late_days = abs(days_diff)
            points = -(late_days * 2)
            log_type = 'return_late'
            reason = f'逾期归还《{loan.book.title}》，逾期{late_days}天'
            fine_amount = late_days * float(loan.fine_daily_rate)
            loan.fine_amount = fine_amount
            loan.status = 'pending_payment'
            loan.fine_paid = False
        
        loan.save()
        
        loan.user.update_credit(points, reason, request.user)
        CreditLog.objects.filter(user=loan.user, reason=reason).update(log_type=log_type)
        
        notified_reservation = Reservation.notify_next_reader(loan.book)
        if notified_reservation:
            messages.info(request, f"已通知预约读者：{notified_reservation.user.username}")
        
        if loan.status == 'pending_payment':
            messages.warning(request, f"图书已归还，但存在逾期罚款 ¥{loan.fine_amount:.2f}。请提醒读者缴费。")
        else:
            messages.success(request, f"图书已成功归还。信用分{points:+d}分，当前信用分：{loan.user.credit_score}")
        
    return redirect('loan_manage')

@login_required
def confirm_payment(request, pk):
    if request.user.role != 'admin':
        return redirect('home')
    loan = get_object_or_404(LoanRecord, pk=pk)
    
    if loan.status == 'pending_payment':
        loan.status = 'returned'
        loan.fine_paid = True
        loan.payment_date = date.today()
        loan.save()
        messages.success(request, f"已确认缴费 ¥{loan.fine_amount:.2f}，借阅记录已完成。")
    
    return redirect('loan_manage')
    
@login_required
def system_settings(request):
    if request.user.role != 'admin':
        return redirect('home')
    
    config = SiteConfig.get_solo()
    borrow_rules = BorrowRule.objects.all()
    active_rule = BorrowRule.get_active_rule()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'update_config':
            config.site_title = request.POST.get('site_title')
            config.maintenance_mode = request.POST.get('maintenance_mode') == 'on'
            daily_fine_rate = request.POST.get('daily_fine_rate', '0.5')
            try:
                config.daily_fine_rate = float(daily_fine_rate)
            except ValueError:
                config.daily_fine_rate = 0.5
            config.save()
            messages.success(request, "系统基本配置已更新。")
        elif action == 'create_category':
            name = request.POST.get('name', '').strip()
            parent_id = request.POST.get('parent_id')
            if name:
                parent = None
                if parent_id:
                    parent = Category.objects.get(pk=parent_id)
                Category.objects.create(name=name, parent=parent)
                messages.success(request, "分类创建成功。")
            else:
                messages.error(request, "分类名称不能为空。")
            return redirect('system_settings')
        elif action == 'edit_category':
            category_id = request.POST.get('category_id')
            name = request.POST.get('name', '').strip()
            if category_id and name:
                category = get_object_or_404(Category, pk=category_id)
                category.name = name
                category.save()
                messages.success(request, "分类已更新。")
            else:
                messages.error(request, "分类名称不能为空。")
            return redirect('system_settings')
        elif action == 'delete_category':
            category_id = request.POST.get('category_id')
            if category_id:
                category = get_object_or_404(Category, pk=category_id)
                if category.can_delete():
                    category.delete()
                    messages.success(request, "分类已删除。")
                else:
                    messages.error(request, "该分类下有图书或子分类，无法删除。")
            return redirect('system_settings')
        elif action == 'create_announcement':
            title = request.POST.get('title')
            content = request.POST.get('content')
            Announcement.objects.create(title=title, content=content)
            messages.success(request, "公告发布成功。")
            return redirect('system_settings')
        elif action == 'create_borrow_rule':
            name = request.POST.get('name')
            max_borrow_days = int(request.POST.get('max_borrow_days', 30))
            max_borrow_quantity = int(request.POST.get('max_borrow_quantity', 0))
            max_daily_requests = int(request.POST.get('max_daily_requests', 0))
            allow_renew = request.POST.get('allow_renew') == 'on'
            max_renew_count = int(request.POST.get('max_renew_count', 1))
            renew_days = int(request.POST.get('renew_days', 15))
            is_active = request.POST.get('is_active') == 'on'
            
            if is_active:
                BorrowRule.objects.update(is_active=False)
            
            BorrowRule.objects.create(
                name=name,
                max_borrow_days=max_borrow_days,
                max_borrow_quantity=max_borrow_quantity,
                max_daily_requests=max_daily_requests,
                allow_renew=allow_renew,
                max_renew_count=max_renew_count,
                renew_days=renew_days,
                is_active=is_active
            )
            messages.success(request, "借阅规则创建成功。")
            return redirect('system_settings')
        elif action == 'edit_borrow_rule':
            rule_id = request.POST.get('rule_id')
            rule = get_object_or_404(BorrowRule, pk=rule_id)
            rule.name = request.POST.get('name')
            rule.max_borrow_days = int(request.POST.get('max_borrow_days', 30))
            rule.max_borrow_quantity = int(request.POST.get('max_borrow_quantity', 0))
            rule.max_daily_requests = int(request.POST.get('max_daily_requests', 0))
            rule.allow_renew = request.POST.get('allow_renew') == 'on'
            rule.max_renew_count = int(request.POST.get('max_renew_count', 1))
            rule.renew_days = int(request.POST.get('renew_days', 15))
            rule.save()
            messages.success(request, "借阅规则已更新。")
            return redirect('system_settings')
        elif action == 'toggle_rule_status':
            rule_id = request.POST.get('rule_id')
            rule = get_object_or_404(BorrowRule, pk=rule_id)
            if rule.is_active:
                rule.is_active = False
                rule.save()
                messages.success(request, "规则已禁用。")
            else:
                BorrowRule.objects.update(is_active=False)
                rule.is_active = True
                rule.save()
                messages.success(request, f"规则「{rule.name}」已设为当前启用规则。")
            return redirect('system_settings')
        elif action == 'delete_borrow_rule':
            rule_id = request.POST.get('rule_id')
            rule = get_object_or_404(BorrowRule, pk=rule_id)
            if rule.is_active:
                messages.error(request, "无法删除当前启用的规则，请先禁用。")
            else:
                rule.delete()
                messages.success(request, "规则已删除。")
            return redirect('system_settings')
            
    announcements = Announcement.objects.all().order_by('-created_at')
    
    credit_logs_all = CreditLog.objects.all().select_related('user', 'operator').order_by('-created_at')
    paginator = Paginator(credit_logs_all, 20)
    page_number = request.GET.get('credit_page')
    credit_logs = paginator.get_page(page_number)
    
    top_categories = Category.objects.filter(parent=None).prefetch_related('children')
    
    return render(request, 'admin/settings.html', {
        'announcements': announcements, 
        'config': config,
        'credit_logs': credit_logs,
        'borrow_rules': borrow_rules,
        'active_rule': active_rule,
        'top_categories': top_categories
    })

@login_required
def announcement_delete(request, pk):
    if request.user.role != 'admin':
        return redirect('home')
    Announcement.objects.filter(pk=pk).delete()
    messages.success(request, "公告已删除")
    return redirect('system_settings')

@login_required
def announcement_create(request):
    return redirect('system_settings')
    
def home(request):
    check_expired_reservations()
    
    announcements = Announcement.objects.filter(is_active=True).order_by('-created_at')[:5]
    latest_books = Book.objects.all().order_by('-created_at')[:8]
    config = SiteConfig.get_solo()
    
    credit_ranking = User.objects.filter(role='reader', is_active=True).order_by('-credit_score')[:10]
    
    user_notifications = []
    if request.user.is_authenticated:
        user_notifications = Announcement.objects.filter(
            is_active=True,
            content__contains=request.user.username
        ).order_by('-created_at')[:3]
    
    return render(request, 'books/home.html', {
        'announcements': announcements,
        'latest_books': latest_books,
        'config': config,
        'credit_ranking': credit_ranking,
        'user_notifications': user_notifications
    })

@login_required
@require_POST
def toggle_favorite(request, pk):
    book = get_object_or_404(Book, pk=pk)
    favorites_list, created = BookList.objects.get_or_create(
        user=request.user,
        name='我的收藏',
        defaults={'description': '默认收藏夹', 'visibility': 'private'}
    )
    
    entry = BookListEntry.objects.filter(book_list=favorites_list, book=book).first()
    if entry:
        entry.delete()
        is_favorited = False
    else:
        BookListEntry.objects.create(book_list=favorites_list, book=book)
        is_favorited = True
    
    favorite_count = BookListEntry.objects.filter(book=book).count()
    
    return JsonResponse({
        'success': True,
        'is_favorited': is_favorited,
        'favorite_count': favorite_count
    })

@login_required
def check_favorite(request, pk):
    book = get_object_or_404(Book, pk=pk)
    favorites_list = BookList.objects.filter(user=request.user, name='我的收藏').first()
    is_favorited = False
    if favorites_list:
        is_favorited = BookListEntry.objects.filter(book_list=favorites_list, book=book).exists()
    
    favorite_count = BookListEntry.objects.filter(book=book).count()
    
    return JsonResponse({
        'is_favorited': is_favorited,
        'favorite_count': favorite_count
    })

@login_required
def my_book_lists(request):
    book_lists = BookList.objects.filter(user=request.user).annotate(
        book_count=Count('entries')
    )
    
    all_favorited_books = Book.objects.filter(
        list_entries__book_list__user=request.user
    ).distinct().annotate(
        avg_rating=Avg('reviews__rating'),
        review_count=Count('reviews')
    )
    
    book_list_assignments = {}
    for book in all_favorited_books:
        book_list_assignments[book.id] = list(
            BookList.objects.filter(
                user=request.user,
                entries__book=book
            ).values_list('id', 'name')
        )
    
    return render(request, 'user/my_book_lists.html', {
        'book_lists': book_lists,
        'all_favorited_books': all_favorited_books,
        'book_list_assignments': book_list_assignments,
        'book_list_assignments_json': json.dumps(book_list_assignments)
    })

@login_required
@require_POST
def book_list_create(request):
    name = request.POST.get('name', '').strip()
    description = request.POST.get('description', '').strip()
    
    if not name:
        messages.error(request, '书单名称不能为空。')
        return redirect('my_book_lists')
    
    if BookList.objects.filter(user=request.user, name=name).exists():
        messages.error(request, '您已存在同名书单。')
        return redirect('my_book_lists')
    
    BookList.objects.create(
        user=request.user,
        name=name,
        description=description
    )
    
    messages.success(request, f'书单"{name}"创建成功！')
    return redirect('my_book_lists')

@login_required
@require_POST
def book_list_edit(request, pk):
    book_list = get_object_or_404(BookList, pk=pk, user=request.user)
    
    name = request.POST.get('name', '').strip()
    description = request.POST.get('description', '').strip()
    visibility = request.POST.get('visibility', 'private')
    
    if not name:
        messages.error(request, '书单名称不能为空。')
        return redirect('my_book_lists')
    
    if name != book_list.name and BookList.objects.filter(user=request.user, name=name).exists():
        messages.error(request, '您已存在同名书单。')
        return redirect('my_book_lists')
    
    book_list.name = name
    book_list.description = description
    book_list.visibility = visibility
    book_list.save()
    
    messages.success(request, f'书单"{name}"更新成功！')
    return redirect('my_book_lists')

@login_required
@require_POST
def book_list_delete(request, pk):
    book_list = get_object_or_404(BookList, pk=pk, user=request.user)
    
    if book_list.name == '我的收藏':
        messages.error(request, '默认收藏夹不能删除。')
        return redirect('my_book_lists')
    
    book_list.delete()
    messages.success(request, f'书单"{book_list.name}"已删除。')
    return redirect('my_book_lists')

@login_required
@require_POST
def book_list_add_book(request, list_pk, book_pk):
    book_list = get_object_or_404(BookList, pk=list_pk, user=request.user)
    book = get_object_or_404(Book, pk=book_pk)
    
    if BookListEntry.objects.filter(book_list=book_list, book=book).exists():
        return JsonResponse({'success': False, 'message': '该书已在此书单中'})
    
    BookListEntry.objects.create(book_list=book_list, book=book)
    return JsonResponse({'success': True})

@login_required
@require_POST
def book_list_remove_book(request, list_pk, book_pk):
    book_list = get_object_or_404(BookList, pk=list_pk, user=request.user)
    entry = get_object_or_404(BookListEntry, book_list=book_list, book_id=book_pk)
    entry.delete()
    return JsonResponse({'success': True})

@login_required
@require_POST
def remove_from_all_lists(request, pk):
    book = get_object_or_404(Book, pk=pk)
    user_lists = BookList.objects.filter(user=request.user)
    BookListEntry.objects.filter(book_list__in=user_lists, book=book).delete()
    return JsonResponse({'success': True})

@login_required
def book_list_share(request, pk):
    book_list = get_object_or_404(BookList, pk=pk, user=request.user)
    
    if book_list.visibility != 'public':
        book_list.visibility = 'public'
        book_list.save()
    
    share_token = book_list.generate_share_token()
    share_url = request.build_absolute_uri(f'/shared-list/{share_token}/')
    
    return JsonResponse({
        'success': True,
        'share_url': share_url,
        'share_token': share_token
    })

def shared_book_list(request, token):
    book_list = get_object_or_404(BookList, share_token=token, visibility='public')
    
    books = Book.objects.filter(
        list_entries__book_list=book_list
    ).annotate(
        avg_rating=Avg('reviews__rating'),
        review_count=Count('reviews')
    ).order_by('-list_entries__added_at')
    
    return render(request, 'books/shared_list.html', {
        'book_list': book_list,
        'books': books,
        'owner': book_list.user
    })

def _parse_date_range(request):
    from datetime import datetime
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    start_date = None
    end_date = None
    try:
        if start_date_str:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        if end_date_str:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        start_date = None
        end_date = None
    return start_date, end_date

def _filter_loans_by_date(start_date, end_date):
    loans = LoanRecord.objects.all()
    if start_date:
        loans = loans.filter(borrow_date__gte=start_date)
    if end_date:
        loans = loans.filter(borrow_date__lte=end_date)
    return loans

@login_required
def dashboard_overview(request):
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    from django.db.models import Sum, Avg, Count
    from datetime import date, timedelta

    start_date, end_date = _parse_date_range(request)

    today = date.today()
    first_day_of_month = today.replace(day=1)

    loans_in_range = _filter_loans_by_date(start_date, end_date)
    returned_loans = loans_in_range.filter(status__in=['returned', 'pending_payment'])

    has_date_range = start_date is not None or end_date is not None

    new_books_qs = Book.objects.all()
    if not has_date_range:
        new_books_qs = new_books_qs.filter(created_at__date__gte=first_day_of_month)
    else:
        if start_date:
            new_books_qs = new_books_qs.filter(created_at__date__gte=start_date)
        if end_date:
            new_books_qs = new_books_qs.filter(created_at__date__lte=end_date)
    new_books_count = new_books_qs.count()

    new_users_qs = User.objects.filter(role='reader')
    if not has_date_range:
        new_users_qs = new_users_qs.filter(created_at__date__gte=first_day_of_month)
    else:
        if start_date:
            new_users_qs = new_users_qs.filter(created_at__date__gte=start_date)
        if end_date:
            new_users_qs = new_users_qs.filter(created_at__date__lte=end_date)
    new_users_count = new_users_qs.count()

    total_stock = Book.objects.aggregate(total=Sum('total_stock'))['total'] or 0
    active_loans_count = LoanRecord.objects.filter(status='borrowed').count()
    borrow_rate = (active_loans_count / total_stock * 100) if total_stock > 0 else 0.0

    avg_borrow_days = 0.0
    durations = []
    for loan in returned_loans:
        if loan.return_date:
            duration = (loan.return_date - loan.borrow_date).days
            if duration > 0:
                durations.append(duration)
    if durations:
        avg_borrow_days = round(sum(durations) / len(durations), 1)

    pending_payments = LoanRecord.objects.filter(status='pending_payment')
    total_pending_fine = sum(loan.calculate_fine() for loan in pending_payments)

    stats = {
        'total_books': Book.objects.count(),
        'total_users': User.objects.count(),
        'active_loans': active_loans_count,
        'pending_requests': LoanRecord.objects.filter(status='pending').count(),
        'new_books_count': new_books_count,
        'new_users_count': new_users_count,
        'borrow_rate': round(borrow_rate, 1),
        'avg_borrow_days': avg_borrow_days,
        'pending_payments': pending_payments.count(),
        'monthly_fine_total': 0,
        'monthly_fine_paid': 0,
        'monthly_fine_unpaid': 0,
    }

    if start_date is None and end_date is None:
        paid_this_month = LoanRecord.objects.filter(
            payment_date__gte=first_day_of_month,
            fine_paid=True
        ).aggregate(total=Sum('fine_amount'))['total'] or 0
        stats['monthly_fine_total'] = float(paid_this_month) + total_pending_fine
        stats['monthly_fine_paid'] = float(paid_this_month)
        stats['monthly_fine_unpaid'] = total_pending_fine
    else:
        paid_in_range = loans_in_range.filter(
            fine_paid=True,
            payment_date__isnull=False
        )
        if start_date:
            paid_in_range = paid_in_range.filter(payment_date__gte=start_date)
        if end_date:
            paid_in_range = paid_in_range.filter(payment_date__lte=end_date)
        paid_total = paid_in_range.aggregate(total=Sum('fine_amount'))['total'] or 0
        unpaid_total = sum(loan.calculate_fine() for loan in loans_in_range.filter(status='pending_payment'))
        stats['monthly_fine_total'] = float(paid_total) + unpaid_total
        stats['monthly_fine_paid'] = float(paid_total)
        stats['monthly_fine_unpaid'] = unpaid_total

    return JsonResponse(stats)

@login_required
def dashboard_category_ranking(request):
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    from django.db.models import Count

    start_date, end_date = _parse_date_range(request)
    loans = _filter_loans_by_date(start_date, end_date)

    category_stats = loans.values(
        'book__category__name',
        'book__category__id'
    ).annotate(
        borrow_count=Count('id')
    ).order_by('-borrow_count')

    categories = []
    for stat in category_stats:
        if stat['book__category__name']:
            categories.append({
                'id': stat['book__category__id'],
                'name': stat['book__category__name'],
                'borrow_count': stat['borrow_count']
            })

    return JsonResponse({'categories': categories})

@login_required
def dashboard_user_ranking(request):
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    from django.db.models import Count

    start_date, end_date = _parse_date_range(request)
    loans = _filter_loans_by_date(start_date, end_date)

    user_stats = loans.values(
        'user__id',
        'user__username'
    ).annotate(
        borrow_count=Count('id')
    ).order_by('-borrow_count')[:10]

    users = []
    for idx, stat in enumerate(user_stats, 1):
        users.append({
            'rank': idx,
            'user_id': stat['user__id'],
            'username': stat['user__username'],
            'borrow_count': stat['borrow_count']
        })

    return JsonResponse({'users': users})

@login_required
def dashboard_loan_details(request):
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    start_date, end_date = _parse_date_range(request)
    loans = _filter_loans_by_date(start_date, end_date).select_related('user', 'book').order_by('-borrow_date')

    details = []
    for loan in loans:
        category_name = loan.book.category.name if loan.book.category else '未分类'
        return_date_str = loan.return_date.strftime('%Y-%m-%d') if loan.return_date else ''
        details.append({
            'id': loan.id,
            'username': loan.user.username,
            'book_title': loan.book.title,
            'category': category_name,
            'borrow_date': loan.borrow_date.strftime('%Y-%m-%d'),
            'due_date': loan.due_date.strftime('%Y-%m-%d'),
            'return_date': return_date_str,
            'status': loan.get_status_display(),
            'fine_amount': float(loan.fine_amount) if loan.fine_amount else 0
        })

    return JsonResponse({'loans': details})

@login_required
def dashboard_chart_data_v2(request):
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    from datetime import timedelta

    start_date, end_date = _parse_date_range(request)

    if start_date and end_date:
        days_total = (end_date - start_date).days + 1
        labels = []
        data = []
        for i in range(days_total):
            target_date = start_date + timedelta(days=i)
            labels.append(target_date.strftime('%m-%d'))
            count = LoanRecord.objects.filter(borrow_date=target_date).count()
            data.append(count)
        return JsonResponse({
            'labels': labels,
            'data': data
        })

    today = timezone.now().date()
    weekly_labels = []
    weekly_data = []
    for i in range(6, -1, -1):
        target_date = today - timedelta(days=i)
        weekly_labels.append(target_date.strftime('%m-%d'))
        count = LoanRecord.objects.filter(borrow_date=target_date).count()
        weekly_data.append(count)

    monthly_labels = ['前四周', '前三周', '前两周', '本周']
    monthly_data = []
    for i in range(3, -1, -1):
        start_d = today - timedelta(days=(i * 7) + 6)
        end_d = today - timedelta(days=i * 7)
        count = LoanRecord.objects.filter(borrow_date__range=[start_d, end_d]).count()
        monthly_data.append(count)

    return JsonResponse({
        'weekly': {
            'labels': weekly_labels,
            'data': weekly_data
        },
        'monthly': {
            'labels': monthly_labels,
            'data': monthly_data
        }
    })

@login_required
def dashboard_export_excel(request):
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.db.models import Sum
    from datetime import date, datetime
    from django.http import HttpResponse
    import io

    start_date, end_date = _parse_date_range(request)
    loans = _filter_loans_by_date(start_date, end_date).select_related('user', 'book').order_by('-borrow_date')

    today = date.today()
    first_day_of_month = today.replace(day=1)

    wb = Workbook()

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4f46e5', end_color='4f46e5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def style_header(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def style_data(ws, start_row, end_row, num_cols):
        for row in range(start_row, end_row + 1):
            for col in range(1, num_cols + 1):
                ws.cell(row=row, column=col).border = thin_border

    ws1 = wb.active
    ws1.title = '概览数据'

    total_stock = Book.objects.aggregate(total=Sum('total_stock'))['total'] or 0
    active_loans_count = LoanRecord.objects.filter(status='borrowed').count()
    borrow_rate = (active_loans_count / total_stock * 100) if total_stock > 0 else 0.0

    returned_loans = loans.filter(status__in=['returned', 'pending_payment'])
    durations = []
    for loan in returned_loans:
        if loan.return_date:
            duration = (loan.return_date - loan.borrow_date).days
            if duration > 0:
                durations.append(duration)
    avg_borrow_days = round(sum(durations) / len(durations), 1) if durations else 0

    ws1['A1'] = '图书馆管理系统 - 数据报表'
    ws1['A1'].font = Font(bold=True, size=16)
    ws1.merge_cells('A1:D1')
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws1['A2'] = f'报表生成时间：{datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws1.merge_cells('A2:D2')
    ws1['A2'].alignment = Alignment(horizontal='center')
    ws1['A2'].font = Font(size=10, color='666666')

    date_range_text = '全部时间'
    if start_date and end_date:
        date_range_text = f'{start_date.strftime("%Y-%m-%d")} 至 {end_date.strftime("%Y-%m-%d")}'
    ws1['A3'] = f'统计范围：{date_range_text}'
    ws1.merge_cells('A3:D3')
    ws1['A3'].alignment = Alignment(horizontal='center')
    ws1['A3'].font = Font(size=10, color='666666')

    ws1.append([])
    ws1.append(['指标', '数值', '指标', '数值'])
    style_header(ws1, 5, 4)

    has_date_range_export = start_date is not None or end_date is not None

    new_books_qs = Book.objects.all()
    if not has_date_range_export:
        new_books_qs = new_books_qs.filter(created_at__date__gte=first_day_of_month)
    else:
        if start_date:
            new_books_qs = new_books_qs.filter(created_at__date__gte=start_date)
        if end_date:
            new_books_qs = new_books_qs.filter(created_at__date__lte=end_date)
    new_books_count = new_books_qs.count()

    new_users_qs = User.objects.filter(role='reader')
    if not has_date_range_export:
        new_users_qs = new_users_qs.filter(created_at__date__gte=first_day_of_month)
    else:
        if start_date:
            new_users_qs = new_users_qs.filter(created_at__date__gte=start_date)
        if end_date:
            new_users_qs = new_users_qs.filter(created_at__date__lte=end_date)
    new_users_count = new_users_qs.count()

    paid_in_range = loans.filter(fine_paid=True, payment_date__isnull=False)
    if start_date:
        paid_in_range = paid_in_range.filter(payment_date__gte=start_date)
    if end_date:
        paid_in_range = paid_in_range.filter(payment_date__lte=end_date)
    paid_total = float(paid_in_range.aggregate(total=Sum('fine_amount'))['total'] or 0)
    unpaid_total = sum(loan.calculate_fine() for loan in loans.filter(status='pending_payment'))

    stats_rows = [
        ['图书总数', Book.objects.count(), '注册用户数', User.objects.count()],
        ['本月新增图书', new_books_count, '本月新注册用户', new_users_count],
        ['借出中图书', active_loans_count, '图书借出率', f'{round(borrow_rate, 1)}%'],
        ['总借阅次数', loans.count(), '平均借阅周期', f'{avg_borrow_days} 天'],
        ['待审核申请', LoanRecord.objects.filter(status='pending').count(), '待缴费笔数', loans.filter(status='pending_payment').count()],
        ['已收罚款', f'¥{paid_total:.2f}', '未收罚款', f'¥{unpaid_total:.2f}'],
    ]

    for row in stats_rows:
        ws1.append(row)

    style_data(ws1, 6, 5 + len(stats_rows), 4)

    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 20

    ws2 = wb.create_sheet('借阅明细')
    headers2 = ['序号', '读者', '图书名称', '分类', '借阅日期', '应还日期', '归还日期', '状态', '罚款金额']
    ws2.append(headers2)
    style_header(ws2, 1, len(headers2))

    for idx, loan in enumerate(loans, 1):
        category_name = loan.book.category.name if loan.book.category else '未分类'
        return_date_str = loan.return_date.strftime('%Y-%m-%d') if loan.return_date else '-'
        fine_str = f'¥{float(loan.fine_amount):.2f}' if loan.fine_amount else '¥0.00'
        ws2.append([
            idx,
            loan.user.username,
            loan.book.title,
            category_name,
            loan.borrow_date.strftime('%Y-%m-%d'),
            loan.due_date.strftime('%Y-%m-%d'),
            return_date_str,
            loan.get_status_display(),
            fine_str
        ])

    if loans.count() > 0:
        style_data(ws2, 2, 1 + loans.count(), len(headers2))

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 30
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 15
    ws2.column_dimensions['F'].width = 15
    ws2.column_dimensions['G'].width = 15
    ws2.column_dimensions['H'].width = 12
    ws2.column_dimensions['I'].width = 12

    ws3 = wb.create_sheet('用户排行')
    headers3 = ['排名', '用户名', '借阅次数']
    ws3.append(headers3)
    style_header(ws3, 1, len(headers3))

    from django.db.models import Count as DBCount
    user_stats = loans.values('user__id', 'user__username').annotate(
        borrow_count=DBCount('id')
    ).order_by('-borrow_count')[:10]

    for idx, stat in enumerate(user_stats, 1):
        ws3.append([idx, stat['user__username'], stat['borrow_count']])

    if user_stats.count() > 0:
        style_data(ws3, 2, 1 + user_stats.count(), len(headers3))

    ws3.column_dimensions['A'].width = 10
    ws3.column_dimensions['B'].width = 25
    ws3.column_dimensions['C'].width = 15

    ws4 = wb.create_sheet('分类统计')
    headers4 = ['排名', '分类名称', '借阅次数', '占比']
    ws4.append(headers4)
    style_header(ws4, 1, len(headers4))

    category_stats = loans.values(
        'book__category__name',
        'book__category__id'
    ).annotate(
        borrow_count=DBCount('id')
    ).order_by('-borrow_count')

    total_loans_count = loans.count()
    for idx, stat in enumerate(category_stats, 1):
        cat_name = stat['book__category__name'] or '未分类'
        percentage = (stat['borrow_count'] / total_loans_count * 100) if total_loans_count > 0 else 0
        ws4.append([idx, cat_name, stat['borrow_count'], f'{round(percentage, 1)}%'])

    if category_stats.count() > 0:
        style_data(ws4, 2, 1 + category_stats.count(), len(headers4))

    ws4.column_dimensions['A'].width = 10
    ws4.column_dimensions['B'].width = 25
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['D'].width = 15

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'图书馆数据报表_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
def book_barcode(request, pk):
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    book = get_object_or_404(Book, pk=pk)
    from .utils import ensure_barcode_exists
    
    barcode_url = ensure_barcode_exists(book)
    
    return JsonResponse({
        'success': True,
        'book_id': book.id,
        'book_title': book.title,
        'isbn': book.isbn,
        'barcode_url': barcode_url
    })


@login_required
def book_barcode_download(request, pk):
    if request.user.role != 'admin':
        return redirect('home')
    
    book = get_object_or_404(Book, pk=pk)
    from .utils import generate_barcode_image, normalize_isbn
    
    filepath, url = generate_barcode_image(book.isbn)
    
    from django.http import FileResponse
    import os
    
    filename = f"{book.title}_{book.isbn}.png"
    response = FileResponse(open(filepath, 'rb'), content_type='image/png')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
@require_POST
def batch_barcode_generate(request):
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    book_ids = request.POST.getlist('book_ids')
    if not book_ids:
        messages.error(request, "请先选择要生成条码的图书。")
        return redirect('book_manage')
    
    books = Book.objects.filter(id__in=book_ids)
    if not books.exists():
        messages.error(request, "未找到选中的图书。")
        return redirect('book_manage')
    
    from .utils import generate_barcode_image
    import zipfile
    from io import BytesIO
    from django.http import HttpResponse
    
    zip_buffer = BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        for book in books:
            if book.isbn:
                filepath, url = generate_barcode_image(book.isbn)
                filename = f"{book.title}_{book.isbn}.png"
                filename = filename.replace('/', '_').replace('\\', '_')
                zip_file.write(filepath, arcname=filename)
    
    zip_buffer.seek(0)
    
    from datetime import datetime
    zip_filename = f"图书条码批量生成_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
    
    response = HttpResponse(
        zip_buffer.getvalue(),
        content_type='application/zip'
    )
    response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'
    return response


@login_required
def scan_borrow_return(request):
    if request.user.role != 'admin':
        return redirect('home')
    
    return render(request, 'admin/scan_borrow_return.html')


@login_required
def scan_lookup(request):
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    isbn = request.GET.get('isbn', '').strip()
    if not isbn:
        return JsonResponse({'success': False, 'message': '请输入ISBN'})
    
    from .utils import normalize_isbn
    normalized_isbn = normalize_isbn(isbn)
    
    book = Book.objects.filter(isbn=isbn).first()
    if not book:
        book = Book.objects.filter(isbn=normalized_isbn).first()
    
    if not book:
        return JsonResponse({'success': False, 'message': '未找到该ISBN对应的图书'})
    
    active_loans = LoanRecord.objects.filter(
        book=book,
        status='borrowed'
    ).select_related('user').order_by('borrow_date')
    
    from .utils import ensure_barcode_exists
    barcode_url = ensure_barcode_exists(book)
    
    book_data = {
        'id': book.id,
        'title': book.title,
        'author': book.author,
        'isbn': book.isbn,
        'stock': book.stock,
        'total_stock': book.total_stock,
        'cover_url': book.cover.url if book.cover else None,
        'barcode_url': barcode_url,
        'category_name': book.category.get_full_name() if book.category else '未分类',
        'borrowed_count': active_loans.count(),
    }
    
    loans_data = []
    for loan in active_loans:
        loans_data.append({
            'id': loan.id,
            'borrow_date': loan.borrow_date.strftime('%Y-%m-%d'),
            'due_date': loan.due_date.strftime('%Y-%m-%d'),
            'is_overdue': loan.is_overdue(),
            'user': {
                'id': loan.user.id,
                'username': loan.user.username,
                'email': loan.user.email,
            }
        })
    
    return JsonResponse({
        'success': True,
        'book': book_data,
        'active_loans': loans_data
    })


@login_required
@require_POST
def scan_return_book(request, loan_id):
    if request.user.role != 'admin':
        return JsonResponse({'error': 'Unauthorized'}, status=403)
    
    loan = get_object_or_404(LoanRecord, pk=loan_id, status='borrowed')
    
    loan.return_date = date.today()
    loan.book.stock += 1
    loan.book.save()
    
    days_diff = (loan.due_date - loan.return_date).days
    if days_diff >= 7:
        points = 3
        log_type = 'return_early'
        reason = f'提前归还《{loan.book.title}》，提前{days_diff}天'
        loan.status = 'returned'
        loan.fine_paid = True
    elif days_diff >= 0:
        points = 1
        log_type = 'return_on_time'
        reason = f'按时归还《{loan.book.title}》'
        loan.status = 'returned'
        loan.fine_paid = True
    else:
        late_days = abs(days_diff)
        points = -(late_days * 2)
        log_type = 'return_late'
        reason = f'逾期归还《{loan.book.title}》，逾期{late_days}天'
        fine_amount = late_days * float(loan.fine_daily_rate)
        loan.fine_amount = fine_amount
        loan.status = 'pending_payment'
        loan.fine_paid = False
    
    loan.save()
    
    loan.user.update_credit(points, reason, request.user)
    from apps.users.models import CreditLog
    CreditLog.objects.filter(user=loan.user, reason=reason).update(log_type=log_type)
    
    Reservation.notify_next_reader(loan.book)
    
    return JsonResponse({
        'success': True,
        'status': loan.status,
        'points': points,
        'reason': reason,
        'fine_amount': float(loan.fine_amount) if loan.fine_amount else 0,
        'message': '归还成功'
    })
